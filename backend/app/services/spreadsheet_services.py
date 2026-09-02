import csv
import io

from openpyxl import load_workbook

from app.schemas.document import DocumentPayload


class SpreadsheetService:
    """
    Service responsible for extracting structured
    content from spreadsheet documents.

    Currently supports:
    - XLSX
    - CSV
    """

    # =========================================================
    # XLSX
    # =========================================================

    @staticmethod
    def extract_xlsx(
        file_bytes: bytes,
        filename: str,
    ) -> DocumentPayload:
        """
        Extract textual content and metadata
        from an Excel workbook.

        Cached formula results are preferred when
        available. If a formula has no cached result,
        the formula itself is preserved in the
        extracted text instead of silently disappearing.
        """

        # -----------------------------------------------------
        # VALUE WORKBOOK
        # -----------------------------------------------------
        #
        # data_only=True:
        # Formula cells return their cached calculated
        # result when one exists.

        value_workbook = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )

        # -----------------------------------------------------
        # FORMULA WORKBOOK
        # -----------------------------------------------------
        #
        # data_only=False:
        # Formula cells retain expressions such as
        # =B12*C12.

        formula_workbook = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=False,
        )

        sections = []

        formula_count = 0
        unresolved_formula_count = 0

        try:
            for sheet_index, value_sheet in enumerate(
                value_workbook.worksheets
            ):

                formula_sheet = (
                    formula_workbook.worksheets[
                        sheet_index
                    ]
                )

                rows = []

                # Iterate through both versions of
                # the worksheet simultaneously.

                value_rows = value_sheet.iter_rows()

                formula_rows = formula_sheet.iter_rows()

                for value_row, formula_row in zip(
                    value_rows,
                    formula_rows,
                ):

                    values = []

                    for value_cell, formula_cell in zip(
                        value_row,
                        formula_row,
                    ):

                        calculated_value = (
                            value_cell.value
                        )

                        original_value = (
                            formula_cell.value
                        )

                        # -------------------------------------
                        # FORMULA CELL
                        # -------------------------------------

                        is_formula = (
                            isinstance(
                                original_value,
                                str,
                            )
                            and original_value.startswith(
                                "="
                            )
                        )

                        if is_formula:
                            formula_count += 1

                            # Cached result exists.
                            if calculated_value is not None:
                                value = calculated_value

                            # No cached result exists.
                            # Preserve formula rather than
                            # returning an empty cell.
                            else:
                                value = original_value

                                unresolved_formula_count += 1

                        # -------------------------------------
                        # NORMAL CELL
                        # -------------------------------------

                        else:
                            value = calculated_value

                            if (
                                value is None
                                and original_value
                                is not None
                            ):
                                value = original_value

                        # -------------------------------------
                        # NORMALIZE VALUE
                        # -------------------------------------

                        if value is None:
                            normalized_value = ""

                        else:
                            normalized_value = str(
                                value
                            ).strip()

                        values.append(
                            normalized_value
                        )

                    # Ignore completely empty rows.
                    if not any(
                        value.strip()
                        for value in values
                    ):
                        continue

                    rows.append(
                        " | ".join(values)
                    )

                # Add sheet only when it contains
                # usable information.
                if rows:

                    section = (
                        f"Sheet: "
                        f"{value_sheet.title}\n\n"
                        + "\n".join(rows)
                    )

                    sections.append(section)

            extracted_text = (
                "\n\n---\n\n".join(
                    sections
                )
            )

            if not extracted_text.strip():
                raise ValueError(
                    "Spreadsheet contains "
                    "no usable data."
                )

            return DocumentPayload(
                filename=filename,
                file_type="xlsx",
                text=extracted_text,
                metadata={
                    "sheet_count": len(
                        value_workbook.sheetnames
                    ),
                    "sheets": list(
                        value_workbook.sheetnames
                    ),
                    "formula_count": (
                        formula_count
                    ),
                    "unresolved_formula_count": (
                        unresolved_formula_count
                    ),
                },
            )

        finally:
            value_workbook.close()

            formula_workbook.close()

    # =========================================================
    # CSV
    # =========================================================

    @staticmethod
    def extract_csv(
        file_bytes: bytes,
        filename: str,
    ) -> DocumentPayload:
        """
        Extract textual content and metadata
        from a CSV file.
        """

        # utf-8-sig handles normal UTF-8 CSV files
        # as well as files containing a UTF-8 BOM.

        text = file_bytes.decode(
            "utf-8-sig",
            errors="replace",
        )

        reader = csv.reader(
            io.StringIO(text)
        )

        rows = []

        max_columns = 0

        for row in reader:

            values = [
                str(value).strip()
                for value in row
            ]

            # Ignore completely empty rows.
            if not any(values):
                continue

            max_columns = max(
                max_columns,
                len(values),
            )

            rows.append(
                " | ".join(values)
            )

        extracted_text = "\n".join(
            rows
        )

        if not extracted_text.strip():
            raise ValueError(
                "CSV contains no usable data."
            )

        return DocumentPayload(
            filename=filename,
            file_type="csv",
            text=extracted_text,
            metadata={
                # First non-empty row is treated
                # as the CSV header.
                "row_count": max(
                    len(rows) - 1,
                    0,
                ),
                "column_count": max_columns,
            },
        )